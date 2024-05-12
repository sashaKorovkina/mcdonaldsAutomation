from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
import time
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

driver = webdriver.Chrome()


def select_from_dropdown(page, dropdown_xpath, option_text, wait=10, sleep=3):
    dropdown = WebDriverWait(page, wait).until(
        EC.element_to_be_clickable((By.XPATH, dropdown_xpath))
    )
    select = Select(dropdown)
    select.select_by_visible_text(option_text)
    time.sleep(sleep)


def driver_page(url):

    return driver


def sender_id(page, button, input, wait=10, sleep=3):
    change_button = WebDriverWait(page, wait).until(
        EC.element_to_be_clickable((By.ID, button))
    )
    change_button.click()
    change_button.clear()
    change_button.send_keys(input)
    time.sleep(sleep)


def clicker_id(page, button, wait=10, sleep=3):
    change_button = WebDriverWait(page, wait).until(
        EC.element_to_be_clickable((By.ID, button))
    )
    change_button.click()
    time.sleep(sleep)


def clicker_xpath(page, button, wait=10, sleep=3):
    change_button = WebDriverWait(page, wait).until(
        EC.element_to_be_clickable((By.XPATH, button))
    )
    change_button.click()
    time.sleep(sleep)


def sender_xpath(page, button, input, wait=10, sleep=3):
    change_button = WebDriverWait(page, wait).until(
        EC.element_to_be_clickable((By.XPATH, button))
    )
    change_button.click()
    change_button.clear()
    change_button.send_keys(input)
    change_button.send_keys(Keys.ENTER)
    time.sleep(sleep)

def login(user_name, password):
    print("Logging in...")
    login_url  = "https://www.mcdstuff.co.uk/portal.php?site=login&page=mcdonaldsUKLogin"
    driver.get(login_url )
    clicker_xpath(driver, "/html/body/portal-ui/div/div[1]/a[1]/span[2]")
    clicker_id(driver, "btnSetPopup")
    login_id = "//div[@class='title' and contains(text(), 'Restaurant Managers & Franchisees')]"
    clicker_xpath(driver, login_id)
    sender_id(driver, "UsernameInputTxtManagers", user_name)
    sender_id(driver, "PasswordInputManagers", password)
    clicker_id(driver, "btnLoginManagers", wait=30)
    hr_page = "/html/body/div[3]/portal-ui/div[3]/a[15]/div"
    clicker_xpath(driver, hr_page, wait=30)
    original_window = driver.current_window_handle
    all_windows = driver.window_handles
    for window in all_windows:
        if window != original_window:
            driver.switch_to.window(window)
    return driver.current_window_handle


def update_employee(employee_id, new_hourly_rate, driver_window_handle):
    driver.switch_to.window(driver_window_handle)
    time.sleep(10)
    print(f"Filtering for employee {employee_id}...")
    filter_employee = '/html/body/div/div[2]/portal-ui/div/portal-ui/div[3]/div[1]/section[1]/input'
    sender_xpath(driver, filter_employee, employee_id, wait=30)
    click_actions = "/html/body/div/div[2]/portal-ui/div/portal-ui/div[3]/div[2]/div[1]/div/div/div/button"
    clicker_xpath(driver, click_actions)
    click_manage_assignments = "/html/body/div/div[2]/portal-ui/div/portal-ui/div[3]/div[2]/div[1]/div/div/div[2]/div/div[1]/a[5]"
    clicker_xpath(driver, click_manage_assignments, wait=30)
    original_window = driver.current_window_handle
    all_windows = driver.window_handles
    for window in all_windows:
        if window != original_window:
            driver.switch_to.window(window)
    time.sleep(10)

    click_change = "/html/body/div[1]/div[2]/portal-ui/div/portal-ui/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[2]/button[1]"
    clicker_xpath(driver, click_change)

    print(f"Managing employee pay...")
    change_type_dropdown = "/html/body/div[3]/div/portal-ui/div/select[1]"
    clicker_xpath(driver, change_type_dropdown)
    change_type = "/html/body/div[3]/div/portal-ui/div/select[1]"
    select_from_dropdown(driver, change_type, "Pay Change")
    change_reason_dropdown = "/html/body/div[3]/div/portal-ui/div/select"
    clicker_xpath(driver, change_reason_dropdown)
    change_reason = "/html/body/div[3]/div/portal-ui/div/select[2]"
    select_from_dropdown(driver, change_reason, "Scale Change")
    as_of_date = "/html/body/div[3]/div/portal-ui/div/div/div/input[2]"
    date_string = '06/05/2024'

    # Convert the string to a datetime object
    # Make sure the format in strptime matches the format of your date_string
    today_string = datetime.datetime.strptime(date_string, '%d/%m/%Y')
    # today_string = date_object.strftime("%d/%m/%Y")
    sender_xpath(driver, as_of_date, date_string)
    change_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, as_of_date))
    )
    change_button.click()
    # change_button.clear()
    change_button.send_keys(Keys.ENTER)
    # time.sleep(3)
    add_change = "/html/body/div[3]/div/portal-ui/div/button[1]"
    add_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, add_change))
    )
    add_button.click()

    print(f"Updating pay to {new_hourly_rate}")
    hourly_rate = "/html/body/div[1]/div[2]/portal-ui/div/portal-ui/div/div[2]/div/div/div[4]/div[8]/div[2]/div/div/div/div/input"

    sender_xpath(driver, hourly_rate, new_hourly_rate)

    change_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, hourly_rate))
    )

    input_element = change_button.find_element(By.XPATH, "./preceding-sibling::input")
    print(input_element)
    if input_element == input:
        print("Submitting new pay details")
        # /html/body/div[8]/div/portal-ui/div/button[1]
        # /html/body/div[1]/div[2]/portal-ui/div/portal-ui/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[1]/button[4]
        # /html/body/div[1]/div[2]/portal-ui/div/portal-ui/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[1]/button[1]
        click_submit = "/html/body/div[1]/div[2]/portal-ui/div/portal-ui/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[1]/button[4]"
        clicker_xpath(driver, click_submit)

        print("New details submitted...")
        # print("Canceling new pay details")
        # click_cancel = "/html/body/div[1]/div[2]/portal-ui/div/portal-ui/div/div[2]/div/div/div[2]/div[2]/div/div[2]/div[1]/button[6]"
        # clicker_xpath(driver, click_cancel)
        print("Close tab in 10 secs...")
        time.sleep(10)
        driver.close()

if __name__ == "__main__":
    errored_employee = []
    file_path = r"C:\Users\sasha\Desktop\McDonalds\Payrates.xlsx"
    sheet_name = "SM £13"
    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl', skiprows=0)
    column_b = df.iloc[:, 1]  # Column B
    column_e = df.iloc[:, 5]  # Column F

    # Combine columns into a 2D array of pairs
    combined = list(zip(column_b, column_e))
    driver_handle = login(user_name="ee949499", password="Cervantes_cag2016")
    # print(combined)
    for employee_id, new_hourly_rate in combined:
        try:
            print(f'{new_hourly_rate}')
            print(f'{employee_id}')
            update_employee(employee_id=employee_id, new_hourly_rate=str(new_hourly_rate),
                            driver_window_handle=driver_handle)
        except Exception as e:
            print(f'Errored employee_id: {employee_id}')
            errored_employee.append(employee_id)

print(errored_employee)

wb = load_workbook(file_path)
ws = wb[sheet_name]

for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=2, max_col=2):
    employee_name = row[0].value  # Assuming column B contains employee names
    if employee_name in errored_employee:
        for cell in row:
            cell.fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

# Save the workbook
wb.save(file_path)
    # update_employee(employee_id=282463, new_hourly_rate="13.50", driver_window_handle=driver_handle)
    # update_employee(employee_id=124441, new_hourly_rate="12.50", driver_window_handle=driver_handle)
    # update_employee(employee_id=153193, new_hourly_rate="12.50", driver_window_handle=driver_handle)
    # update_employee(employee_id=210458, new_hourly_rate="12.50", driver_window_handle=driver_handle)
